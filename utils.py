import os
from pypdf import PdfReader
from docx import Document
from pptx import Presentation
from openpyxl import load_workbook
from PIL import Image
import math

def get_page_count(file_path):
    """
    Returns the number of pages in a file.
    Supports: .pdf, .docx, .pptx, .xlsx, .xls, and image files (.jpg, .jpeg, .png, .bmp, .gif, .tiff)
    Returns 1 for other file types or if counting fails.
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if ext == '.pdf':
            reader = PdfReader(file_path)
            return len(reader.pages)
        
        elif ext == '.docx':
            # Count pages by estimating from content
            doc = Document(file_path)
            total_chars = sum(len(para.text) for para in doc.paragraphs)
            # Estimate: ~3000 characters per page (rough approximation)
            estimated_pages = max(1, math.ceil(total_chars / 3000))
            return estimated_pages
        
        elif ext in ['.pptx', '.ppt']:
            prs = Presentation(file_path)
            return len(prs.slides)
        
        elif ext in ['.xlsx', '.xls']:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            # Count non-empty sheets
            sheet_count = len([sheet for sheet in wb.worksheets if sheet.max_row > 0])
            return max(1, sheet_count)
        
        elif ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif']:
            # Images count as 1 page
            # Verify it's a valid image file
            with Image.open(file_path) as img:
                img.verify()
            return 1
        
        else:
            return 1
    
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return 1

def format_currency(amount):
    return f"{amount:,.0f} VND".replace(",", ".")
